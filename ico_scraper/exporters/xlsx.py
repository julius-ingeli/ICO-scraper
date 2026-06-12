from __future__ import annotations

import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D


TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "template.xlsx"
SOURCE_URL_FIELD = "__source_url"
GENERAL_INFO_FIELD = "__general_info"

# Excel stores column widths in character units. This maps requested pixel widths
# closely enough for the default Calibri-based workbook rendering.
EXCEL_WIDTH_PADDING_PX = 5
EXCEL_WIDTH_CHAR_PX = 7

OUTPUT_COLUMN_WIDTHS_PX = {
    "Prehľad": {"C": 400},
    "ORSR": {"A": 400, "D": 400},
    "ORSR - Podnikanie": {"A": 600},
    "ORSR - Právne skutočnosti": {"A": 1200},
    "RPVS": {"D": 650},
    "Finstat": {"D": 650},
    "CRZ": {"B": 400},
}

LABELS = {
    "obchodne_meno": "Obchodné meno",
    "sidlo": "Sídlo",
    "den_zapisu": "Deň zápisu",
    "pravna_forma": "Právna forma",
    "vyska_zakladneho_imania": "Výška základného imania",
    "typ_organu": "Typ orgánu",
    "statutarny_organ": "Štatutárny orgán",
    "pravny_predchodca": "Právny predchodca",
    "konanie_menom_spolocnosti": "Konanie menom spoločnosti",
    "dozorna_rada": "Dozorná rada",
    "akcionar": "Akcionár",
    "dalsie_pravne_skutocnosti": "Ďalšie právne skutočnosti",
    "datum_aktualizacie_dat": "Dátum aktualizácie dát",
    "partner_verejneho_sektora": "Partner verejného sektora",
    "opravnena_osoba": "Oprávnená osoba",
    "konecni_uzivatelia_vyhod": "Koneční užívatelia výhod",
    "oznamenie_o_overeni_konecnych_uzivatelov_vyhod": "Oznámenie o overení konečných užívateľov výhod",
    "uctovne_zavierky": "Účtovné závierky",
    "vyrocne_spravy": "Výročné správy",
    "nazov_zmluvy": "Názov zmluvy",
    "cislo_zmluvy": "Číslo zmluvy",
    "dodavatel": "Dodávateľ",
    "odberatel": "Odberateľ",
    "vyhladavanie_podla": "Vyhľadávanie podľa",
    "hladany_dodavatel": "Hľadaný dodávateľ",
}


def export_filename(data: dict[str, Any]) -> str:
    ico = str(data.get("ico") or "export").strip() or "export"
    ico = re.sub(r"[^0-9A-Za-z_-]", "", ico)
    return f"export_{ico}.xlsx"


def label_for(key: Any) -> str:
    text = str(key)
    return LABELS.get(text, text.replace("_", " ").capitalize())


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        if value.startswith("data:image/"):
            return "[obrázok grafu vynechaný]"
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, dict):
        if set(value.keys()) == {"message"}:
            return value.get("message") or ""
        return "\n".join(f"{label_for(k)}: {value_to_text(v)}" for k, v in value.items() if k != SOURCE_URL_FIELD)
    if isinstance(value, list):
        return "\n".join(value_to_text(item) for item in value)
    return str(value)


def fill_value(ws, cell_ref: str, value: Any, hyperlink: str | None = None) -> None:
    cell = ws[cell_ref]
    cell.value = value_to_text(value)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.style = "Hyperlink"


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 7) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def ensure_rows(ws, start_row: int, reserved_rows: int, required_rows: int, max_col: int = 7) -> None:
    if required_rows <= reserved_rows:
        return
    insert_at = start_row + reserved_rows
    extra = required_rows - reserved_rows
    ws.insert_rows(insert_at, extra)
    for row in range(insert_at, insert_at + extra):
        copy_row_style(ws, start_row, row, max_col=max_col)


def clear_row_values(ws, row: int, max_col: int = 7) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def excel_width_from_pixels(px: int | float) -> float:
    return max(0, (float(px) - EXCEL_WIDTH_PADDING_PX) / EXCEL_WIDTH_CHAR_PX)


def apply_output_column_widths(wb) -> None:
    for sheet_name, columns in OUTPUT_COLUMN_WIDTHS_PX.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for column, px in columns.items():
            ws.column_dimensions[column].width = excel_width_from_pixels(px)
            ws.column_dimensions[column].hidden = False


def table_for_header_row(ws, header_row: int):
    for table in ws.tables.values():
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row == header_row:
            return table, min_col, max_col
    return None, None, None


def update_table_definition(ws, header_row: int, data_start_row: int, row_count: int, headers: list[str]) -> None:
    table, min_col, max_col = table_for_header_row(ws, header_row)
    if table is None:
        return
    last_row = max(data_start_row, data_start_row + max(row_count, 1) - 1)
    max_col = min_col + len(headers) - 1 if min_col else max_col
    table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{last_row}"
    if table.autoFilter:
        table.autoFilter.ref = table.ref
    for index, header in enumerate(headers):
        if index < len(table.tableColumns):
            table.tableColumns[index].name = str(header)


def write_pairs(ws, start_row: int, pairs: list[tuple[Any, Any]], reserved_rows: int = 1) -> None:
    rows = pairs or [("", "")]
    ensure_rows(ws, start_row, reserved_rows, len(rows))
    for offset, (key, value) in enumerate(rows):
        row = start_row + offset
        clear_row_values(ws, row)
        ws.cell(row, 1).value = label_for(key) if key else ""
        ws.cell(row, 4).value = value_to_text(value)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")


def write_lines(ws, start_row: int, lines: list[Any], reserved_rows: int = 1) -> None:
    rows = lines or [""]
    ensure_rows(ws, start_row, reserved_rows, len(rows))
    for offset, value in enumerate(rows):
        row = start_row + offset
        clear_row_values(ws, row)
        ws.cell(row, 1).value = value_to_text(value)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")


def dict_pairs(data: dict[str, Any], skip: set[str] | None = None) -> list[tuple[str, Any]]:
    skip = skip or set()
    return [(key, value) for key, value in data.items() if key not in skip and key != SOURCE_URL_FIELD]


def row_text(row: Any) -> str:
    if isinstance(row, list):
        return "\n".join(value_to_text(item) for item in row)
    return value_to_text(row)


def write_record_lines(ws, start_row: int, records: list[Any], reserved_rows: int = 1) -> None:
    write_lines(ws, start_row, [row_text(record) for record in records], reserved_rows=reserved_rows)


def write_table(
    ws,
    start_row: int,
    records: list[dict[str, Any]],
    columns: list[tuple[str, str]],
    reserved_rows: int,
    hyperlink_getter=None,
) -> None:
    rows = records or []
    ensure_rows(ws, start_row, reserved_rows, max(len(rows), 1))
    update_table_definition(ws, start_row - 1, start_row, len(rows), [label for _, label in columns])
    if not rows:
        clear_row_values(ws, start_row)
        ws.cell(start_row, 1).value = ""
        return
    for offset, record in enumerate(rows):
        row_num = start_row + offset
        clear_row_values(ws, row_num)
        for col_index, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row_num, col_index)
            cell.value = value_to_text(record.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            hyperlink = hyperlink_getter(record, key) if hyperlink_getter else None
            if hyperlink:
                cell.hyperlink = hyperlink
                cell.style = "Hyperlink"


def find_row_with_value(ws, value: str) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == value:
                return cell.row
    return None


def clean_labeled_value(value: Any, labels: tuple[str, ...]) -> str:
    text = value_to_text(value).strip()
    for label in labels:
        text = re.sub(rf"^\s*{re.escape(label)}\s*:?\s*", "", text, flags=re.IGNORECASE)
    return text.strip()


def clean_od_value(value: Any) -> str:
    text = clean_labeled_value(value, ("Od", "od"))
    match = re.search(r"\(?\s*od\s*:?\s*([^)]*)\)?", text, flags=re.IGNORECASE)
    if match:
        text = match.group(1)
    return text.strip().strip("()")


def parse_orsr_stock_row(row: Any) -> dict[str, str]:
    values = row if isinstance(row, list) else [row]
    parsed = {"pocet": "", "druh": "", "podoba": "", "forma": "", "menovita_hodnota": "", "od": ""}
    unlabeled = []

    for item in values:
        text = value_to_text(item).strip()
        if not text:
            continue
        lowered = text.casefold()
        if lowered.startswith("počet") or lowered.startswith("pocet"):
            parsed["pocet"] = clean_labeled_value(text, ("Počet", "Pocet"))
        elif lowered.startswith("druh"):
            parsed["druh"] = clean_labeled_value(text, ("Druh",))
        elif lowered.startswith("podoba"):
            parsed["podoba"] = clean_labeled_value(text, ("Podoba",))
        elif lowered.startswith("forma"):
            parsed["forma"] = clean_labeled_value(text, ("Forma",))
        elif lowered.startswith("menovitá hodnota") or lowered.startswith("menovita hodnota"):
            parsed["menovita_hodnota"] = clean_labeled_value(text, ("Menovitá hodnota", "Menovita hodnota"))
        elif re.search(r"\(?\s*od\s*:", lowered) or lowered.startswith("od"):
            parsed["od"] = clean_od_value(text)
        else:
            unlabeled.append(text)

    ordered_keys = ["pocet", "druh", "podoba", "forma", "menovita_hodnota", "od"]
    for key, value in zip([key for key in ordered_keys if not parsed[key]], unlabeled):
        parsed[key] = clean_od_value(value) if key == "od" else value

    return parsed


def write_orsr_stocks_table(ws, start_row: int, stocks: list[Any], reserved_rows: int = 1) -> None:
    rows = [parse_orsr_stock_row(stock) for stock in (stocks or [])]
    ensure_rows(ws, start_row, reserved_rows, max(len(rows), 1))
    if not rows:
        clear_row_values(ws, start_row)
        return
    columns = ["pocet", "druh", "podoba", "forma", "menovita_hodnota", "od"]
    update_table_definition(ws, start_row - 1, start_row, len(rows), ["Počet", "Druh", "Podoba", "Forma", "Menovitá hodnota", "Od"])
    for offset, record in enumerate(rows):
        row_num = start_row + offset
        clear_row_values(ws, row_num)
        for col_index, key in enumerate(columns, start=1):
            ws.cell(row_num, col_index).value = record.get(key, "")
            ws.cell(row_num, col_index).alignment = Alignment(wrap_text=True, vertical="top")


def get_ci(data: dict[str, Any], *keys: str) -> Any:
    if not isinstance(data, dict):
        return ""
    normalized = {str(key).casefold(): value for key, value in data.items()}
    for key in keys:
        value = normalized.get(key.casefold())
        if value not in (None, ""):
            return value
    return ""


def first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def first_document_url(record: dict[str, Any]) -> str:
    documents = record.get("dokumenty") if isinstance(record, dict) else None
    if not isinstance(documents, list):
        return ""
    for document in documents:
        if isinstance(document, dict) and document.get("url"):
            return str(document["url"])
    return ""


def ruz_table_hyperlink(record: dict[str, Any], key: str) -> str:
    if key == "Obdobie":
        return first_document_url(record)
    return ""


def crz_table_hyperlink(record: dict[str, Any], key: str) -> str:
    if key == "nazov_zmluvy" and isinstance(record, dict):
        return str(record.get("link") or "")
    return ""


def write_overview(wb, data: dict[str, Any]) -> None:
    ws = wb["Prehľad"]
    orsr = data.get("orsr", {}) if isinstance(data.get("orsr"), dict) else {}
    finstat = data.get(GENERAL_INFO_FIELD) if isinstance(data.get(GENERAL_INFO_FIELD), dict) else data.get("finstat", {})
    finstat_basic = finstat.get("zakladne_udaje", {}) if isinstance(finstat, dict) else {}
    fill_value(ws, "C5", data.get("ico", ""))
    fill_value(ws, "C6", first_value(orsr.get("obchodne_meno"), finstat_basic.get("Obchodné meno"), finstat_basic.get("Názov")))
    fill_value(ws, "C7", first_value(orsr.get("sidlo"), finstat_basic.get("Sídlo"), finstat_basic.get("Adresa")))
    fill_value(ws, "C8", first_value(orsr.get("den_zapisu"), finstat_basic.get("Dátum vzniku")))
    ws["A12"].value = None
    apply_overview_reference_layout(ws)
    write_finstat_charts(ws, finstat.get("grafy", []) if isinstance(finstat, dict) else [], start_row=12)


def write_orsr(wb, data: dict[str, Any]) -> None:
    orsr = data.get("orsr", {}) if isinstance(data.get("orsr"), dict) else {}
    ws = wb["ORSR"]
    fixed = {
        "D2": orsr.get("obchodne_meno"),
        "D3": orsr.get("sidlo"),
        "D4": orsr.get("den_zapisu"),
        "D5": orsr.get("pravna_forma"),
        "D6": orsr.get("vyska_zakladneho_imania"),
        "D7": orsr.get("typ_organu"),
        "A23": orsr.get("datum_aktualizacie_dat"),
    }
    for cell, value in fixed.items():
        fill_value(ws, cell, value)
    write_record_lines(ws, 9, orsr.get("statutarny_organ", []), reserved_rows=7)
    write_record_lines(ws, 17, orsr.get("konanie_menom_spolocnosti", []), reserved_rows=1)
    write_record_lines(ws, 19, orsr.get("dozorna_rada", []), reserved_rows=3)
    fill_value(ws, "A25", orsr.get(SOURCE_URL_FIELD, ""), hyperlink=orsr.get(SOURCE_URL_FIELD, ""))

    ws = wb["ORSR - Akcie"]
    write_record_lines(ws, 3, orsr.get("akcionar", []), reserved_rows=1)
    stock_title_row = find_row_with_value(ws, "Akcie") or 4
    stock_header_row = stock_title_row + 1
    headers = ["Počet", "Druh", "Podoba", "Forma", "Menovitá hodnota", "Od"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(stock_header_row, col_index).value = header
    write_orsr_stocks_table(ws, stock_header_row + 1, orsr.get("akcie", []), reserved_rows=1)

    ws = wb["ORSR - Podnikanie"]
    write_lines(ws, 2, orsr.get("predmet_podnikania", []), reserved_rows=1)

    ws = wb["ORSR - Právne skutočnosti"]
    lines = []
    lines.extend(row_text(row) for row in orsr.get("pravny_predchodca", []))
    lines.extend(row_text(row) for row in orsr.get("dalsie_pravne_skutocnosti", []))
    write_lines(ws, 2, lines, reserved_rows=1)


def write_rpvs(wb, data: dict[str, Any]) -> None:
    rpvs = data.get("rpvs", {}) if isinstance(data.get("rpvs"), dict) else {}
    ws = wb["RPVS"]
    partner = rpvs.get("partner_verejneho_sektora", {}) if isinstance(rpvs.get("partner_verejneho_sektora"), dict) else {}
    mapping = {
        "D2": get_ci(partner, "Obchodné meno", "Názov", "Meno / názov"),
        "D3": get_ci(partner, "IČO"),
        "D4": get_ci(partner, "Právna forma"),
        "D5": get_ci(partner, "Adresa", "Sídlo"),
        "D6": get_ci(partner, "Dátum zápisu"),
        "D7": get_ci(partner, "Dátum výmazu"),
        "D8": get_ci(partner, "Číslo vložky"),
    }
    for cell, value in mapping.items():
        fill_value(ws, cell, value)
    opravnena_heading_row = find_row_with_value(ws, "Oprávnená osoba") or 9
    opravnena_pairs = dict_pairs(rpvs.get("opravnena_osoba", {}))
    opravnena_start_row = opravnena_heading_row + 1
    write_pairs(ws, opravnena_start_row, opravnena_pairs, reserved_rows=1)

    spacer_row = opravnena_start_row + max(len(opravnena_pairs), 1)
    ws.insert_rows(spacer_row, 1)
    clear_row_values(ws, spacer_row)

    kuvy_heading_row = find_row_with_value(ws, "Koneční užívatelia výhod") or 11
    kuvy_start_row = kuvy_heading_row + 2
    oznamenie_heading_row = find_row_with_value(ws, "Oznámenie o overení konečných užívateľov výhod") or 19
    kuvy_reserved_rows = max(1, oznamenie_heading_row - kuvy_start_row)
    write_table(ws, kuvy_start_row, rpvs.get("konecni_uzivatelia_vyhod", []), [
        ("Meno a priezvisko", "Meno a priezvisko"),
        ("Dátum narodenia", "Dátum narodenia"),
        ("Štátna príslušnosť", "Štátna príslušnosť"),
        ("Adresa", "Adresa"),
        ("Verejný funkcionár", "Verejný funkcionár"),
    ], reserved_rows=kuvy_reserved_rows)

    oznamenie_heading_row = find_row_with_value(ws, "Oznámenie o overení konečných užívateľov výhod") or oznamenie_heading_row
    zdroj_heading_row = find_row_with_value(ws, "Zdroj") or 25
    oznamenie_reserved_rows = max(1, zdroj_heading_row - (oznamenie_heading_row + 1))
    write_pairs(
        ws,
        oznamenie_heading_row + 1,
        dict_pairs(rpvs.get("oznamenie_o_overeni_konecnych_uzivatelov_vyhod", {})),
        reserved_rows=oznamenie_reserved_rows,
    )

    zdroj_heading_row = find_row_with_value(ws, "Zdroj") or zdroj_heading_row
    fill_value(ws, f"A{zdroj_heading_row + 1}", rpvs.get(SOURCE_URL_FIELD, ""), hyperlink=rpvs.get(SOURCE_URL_FIELD, ""))


def write_finstat(wb, data: dict[str, Any]) -> None:
    finstat = data.get("finstat", {}) if isinstance(data.get("finstat"), dict) else {}
    ws = wb["Finstat"]
    basic = finstat.get("zakladne_udaje", {}) if isinstance(finstat.get("zakladne_udaje"), dict) else {}
    financial = finstat.get("financne_ukazovatele", {}) if isinstance(finstat.get("financne_ukazovatele"), dict) else {}
    important = finstat.get("dolezite_ukazovatele", {}) if isinstance(finstat.get("dolezite_ukazovatele"), dict) else {}
    for row in range(3, 11):
        fill_value(ws, f"D{row}", get_ci(basic, str(ws[f"A{row}"].value or "")))
    for row in range(12, 19):
        fill_value(ws, f"D{row}", get_ci(financial, str(ws[f"A{row}"].value or "")))
    for row in range(20, 29):
        label = str(ws[f"A{row}"].value or "").replace(";", "")
        fill_value(ws, f"D{row}", get_ci(important, label))
    fill_value(ws, "A31", finstat.get(SOURCE_URL_FIELD, ""), hyperlink=finstat.get(SOURCE_URL_FIELD, ""))


def write_ruz(wb, data: dict[str, Any]) -> None:
    ruz = data.get("ruz", {}) if isinstance(data.get("ruz"), dict) else {}
    ws = wb["RÚZ"]
    for row in range(2, 14):
        fill_value(ws, f"D{row}", get_ci(ruz, str(ws[f"A{row}"].value or "")))
    statements = ruz.get("uctovne_zavierky") or ruz.get("zaznamy") or []
    annual_reports = ruz.get("vyrocne_spravy") or []
    statement_type = "Typ závierky"
    annual_type = "Typ výročnej správy"
    write_table(ws, 16, statements, [
        ("Obdobie", "Obdobie"), (statement_type, statement_type), ("Predložená dňa", "Predložená dňa"),
        ("Zostavená dňa", "Zostavená dňa"), ("Schválená dňa", "Schválená dňa"), ("SA uložená dňa", "SA uložená dňa"),
    ], reserved_rows=13, hyperlink_getter=ruz_table_hyperlink)
    write_table(ws, 31, annual_reports, [
        ("Obdobie", "Obdobie"), (annual_type, annual_type), ("Predložená dňa", "Predložená dňa"),
        ("Zostavená dňa", "Zostavená dňa"), ("Schválená dňa", "Schválená dňa"), ("SA uložená dňa", "SA uložená dňa"),
    ], reserved_rows=16, hyperlink_getter=ruz_table_hyperlink)
    fill_value(ws, "A48", ruz.get(SOURCE_URL_FIELD, ""), hyperlink=ruz.get(SOURCE_URL_FIELD, ""))


def write_crz(wb, data: dict[str, Any]) -> None:
    crz = data.get("crz", {}) if isinstance(data.get("crz"), dict) else {}
    ws = wb["CRZ"]
    write_table(ws, 3, crz.get("zmluvy", []), [
        ("datum", "Dátum"), ("nazov_zmluvy", "Zmluva"), ("cena", "Cena"),
        ("dodavatel", "Dodávateľ"), ("odberatel", "Odberateľ"),
    ], reserved_rows=11, hyperlink_getter=crz_table_hyperlink)
    fill_value(ws, "A15", crz.get(SOURCE_URL_FIELD, ""), hyperlink=crz.get(SOURCE_URL_FIELD, ""))


def numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    match = re.search(r"-?\d[\d ]*(?:[,.]\d+)?", value)
    if not match:
        return None
    number = float(match.group(0).replace(" ", "").replace(",", "."))
    lowered = value.lower()
    if "mld" in lowered:
        number *= 1_000_000_000
    elif "mil" in lowered:
        number *= 1_000_000
    elif "tis" in lowered:
        number *= 1_000
    return number



OVERVIEW_CHART_LAYOUT = [
    {"title_cell": "A12", "anchor": "A13", "width": 18.0, "height": 8.3, "col_off": 0, "row_off": 0},
    {"title_cell": "L12", "anchor": "K13", "width": 18.0, "height": 8.3, "col_off": 596348, "row_off": 26504},
    {"title_cell": "A28", "anchor": "A29", "width": 26.0, "height": 11.0, "col_off": 0, "row_off": 0},
    {"title_cell": "A51", "anchor": "A52", "width": 6185646 / 360000, "height": 4159624 / 360000, "col_off": 0, "row_off": 0},
    {"title_cell": "P51", "anchor": "O52", "width": 5867400 / 360000, "height": 4095751 / 360000, "col_off": 685028, "row_off": 25483},
]

OVERVIEW_CHART_LAYOUT_BY_TITLE = {
    "zisk": OVERVIEW_CHART_LAYOUT[0],
    "tržby": OVERVIEW_CHART_LAYOUT[1],
    "trzby": OVERVIEW_CHART_LAYOUT[1],
    "celkové výnosy": OVERVIEW_CHART_LAYOUT[2],
    "celkove vynosy": OVERVIEW_CHART_LAYOUT[2],
    "aktíva": OVERVIEW_CHART_LAYOUT[3],
    "aktiva": OVERVIEW_CHART_LAYOUT[3],
    "pasíva": OVERVIEW_CHART_LAYOUT[4],
    "pasiva": OVERVIEW_CHART_LAYOUT[4],
}


def normalized_chart_title(title: Any) -> str:
    return str(title or "").strip().casefold()


def overview_chart_layout(title: Any, rendered_index: int) -> dict[str, Any]:
    return OVERVIEW_CHART_LAYOUT_BY_TITLE.get(normalized_chart_title(title), OVERVIEW_CHART_LAYOUT[min(rendered_index, len(OVERVIEW_CHART_LAYOUT) - 1)])


def apply_overview_reference_layout(ws) -> None:
    for col_index in range(1, 17):
        letter = get_column_letter(col_index)
        ws.column_dimensions[letter].width = 10 if letter == "A" else 13
        ws.column_dimensions[letter].hidden = False

    for row in (6, 8):
        ws.row_dimensions[row].height = 57.6
    ws.row_dimensions[7].height = 144

    for start, end in ((13, 29), (31, 47), (49, 70), (73, 89)):
        for row in range(start, end + 1):
            ws.row_dimensions[row].height = 18


def excel_color(color: Any, fallback: str = "52606D") -> str:
    value = str(color or fallback).strip()
    if value.startswith("rgb"):
        nums = [int(float(num)) for num in re.findall(r"\d+(?:\.\d+)?", value)[:3]]
        if len(nums) == 3:
            return "".join(f"{num:02X}" for num in nums)
    value = value.lstrip("#")
    if len(value) == 3:
        value = "".join(ch * 2 for ch in value)
    if re.fullmatch(r"[0-9A-Fa-f]{6}", value):
        return value.upper()
    return fallback


def chart_value(item: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        value = item.get(key)
        if isinstance(value, (int, float)):
            return float(value)
        parsed = numeric(value)
        if parsed is not None:
            return parsed
    return None


def set_chart_size(chart, layout: dict[str, Any]) -> None:
    chart.width = float(layout.get("width") or 18)
    chart.height = float(layout.get("height") or 8.3)


def chart_size_emu(layout: dict[str, Any]) -> tuple[int, int]:
    return int(round(float(layout.get("width") or 18) * 360000)), int(round(float(layout.get("height") or 8.3) * 360000))


def add_chart_with_layout(ws, chart, layout: dict[str, Any]) -> None:
    row, col = coordinate_to_tuple(str(layout["anchor"]))
    cx, cy = chart_size_emu(layout)
    chart.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=col - 1,
            row=row - 1,
            colOff=int(layout.get("col_off") or 0),
            rowOff=int(layout.get("row_off") or 0),
        ),
        ext=XDRPositiveSize2D(cx=cx, cy=cy),
    )
    ws.add_chart(chart)


def prepare_chart_data_area(ws, start_row: int, start_col: int, width: int, height: int) -> None:
    for col in range(start_col, start_col + width):
        ws.column_dimensions[get_column_letter(col)].hidden = False
    for row in range(start_row, start_row + height):
        for col in range(start_col, start_col + width):
            ws.cell(row, col).value = None


def configure_chart(chart, title: str, layout: dict[str, Any]) -> None:
    chart.title = title
    chart.style = 13
    chart.legend.position = "r"
    chart.visible_cells_only = False
    set_chart_size(chart, layout)


def write_line_chart(ws, title: str, chart_data: dict[str, Any], layout: dict[str, Any], data_row: int, data_col: int) -> bool:
    points = [point for point in chart_data.get("body") or [] if isinstance(point, dict)]
    rows = []
    for index, point in enumerate(points, start=1):
        value = chart_value(point, "value", "value_label")
        if value is None:
            continue
        rows.append((str(point.get("rok") or index), value))
    if len(rows) < 2:
        return False

    prepare_chart_data_area(ws, data_row, data_col, 2, len(rows) + 1)
    ws.cell(data_row, data_col).value = "Rok"
    ws.cell(data_row, data_col + 1).value = title
    for offset, (year, value) in enumerate(rows, start=1):
        ws.cell(data_row + offset, data_col).value = year
        ws.cell(data_row + offset, data_col + 1).value = value

    chart = LineChart()
    configure_chart(chart, title, layout)
    chart.y_axis.majorGridlines = None
    chart.x_axis.title = "Rok"
    chart.y_axis.numFmt = '# ##0'
    data = Reference(ws, min_col=data_col + 1, min_row=data_row, max_row=data_row + len(rows))
    categories = Reference(ws, min_col=data_col, min_row=data_row + 1, max_row=data_row + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    if chart.series:
        series = chart.series[0]
        series.graphicalProperties.line.solidFill = "0F766E"
        series.graphicalProperties.line.width = 25000
        series.marker.symbol = "circle"
        series.marker.size = 5
        series.marker.graphicalProperties.solidFill = "FFFFFF"
        series.marker.graphicalProperties.line.solidFill = "0F766E"
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    add_chart_with_layout(ws, chart, layout)
    return True


def stacked_segment_value(row: dict[str, Any], segment: dict[str, Any]) -> float:
    value = chart_value(segment, "value", "value_label")
    if value is not None:
        return value
    total = chart_value(row, "total_value", "total_label")
    height = chart_value(segment, "height_svg") or 0
    stack_height = chart_value(row, "stack_height_svg") or 0
    if total is not None and stack_height > 0:
        return total * height / stack_height
    return height


def write_bar_chart(ws, title: str, chart_data: dict[str, Any], layout: dict[str, Any], data_row: int, data_col: int) -> bool:
    rows = [row for row in chart_data.get("body") or [] if isinstance(row, dict)]
    legend = [item for item in chart_data.get("legend") or [] if isinstance(item, dict)]
    if not rows or not legend:
        return False

    prepare_chart_data_area(ws, data_row, data_col, len(legend) + 1, len(rows) + 1)
    ws.cell(data_row, data_col).value = "Rok"
    for offset, item in enumerate(legend, start=1):
        ws.cell(data_row, data_col + offset).value = str(item.get("label") or f"Séria {offset}")

    for row_offset, row in enumerate(rows, start=1):
        ws.cell(data_row + row_offset, data_col).value = str(row.get("rok") or row_offset)
        segments = row.get("segments") or []
        for col_offset, legend_item in enumerate(legend, start=1):
            legend_label = str(legend_item.get("label") or "")
            segment = next((candidate for candidate in segments if str(candidate.get("label") or "") == legend_label), None)
            ws.cell(data_row + row_offset, data_col + col_offset).value = stacked_segment_value(row, segment) if segment else 0

    chart = BarChart()
    configure_chart(chart, title, layout)
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.x_axis.title = "Rok"
    chart.y_axis.numFmt = '# ##0'
    data = Reference(ws, min_col=data_col + 1, max_col=data_col + len(legend), min_row=data_row, max_row=data_row + len(rows))
    categories = Reference(ws, min_col=data_col, min_row=data_row + 1, max_row=data_row + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    for index, series in enumerate(chart.series):
        color = excel_color(legend[index].get("color") if index < len(legend) else None)
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color
    add_chart_with_layout(ws, chart, layout)
    return True


def write_pie_chart(ws, title: str, chart_data: dict[str, Any], layout: dict[str, Any], data_row: int, data_col: int) -> bool:
    slices = [item for item in chart_data.get("body") or [] if isinstance(item, dict)]
    rows = []
    for index, item in enumerate(slices, start=1):
        value = chart_value(item, "value", "value_label")
        if value is not None and value > 0:
            rows.append((str(item.get("label") or f"Séria {index}"), value, item.get("color")))
    if not rows:
        return False

    prepare_chart_data_area(ws, data_row, data_col, 2, len(rows) + 1)
    ws.cell(data_row, data_col).value = "Položka"
    ws.cell(data_row, data_col + 1).value = title
    for offset, (label, value, _) in enumerate(rows, start=1):
        ws.cell(data_row + offset, data_col).value = label
        ws.cell(data_row + offset, data_col + 1).value = value

    chart = PieChart()
    configure_chart(chart, title, layout)
    data = Reference(ws, min_col=data_col + 1, min_row=data_row, max_row=data_row + len(rows))
    categories = Reference(ws, min_col=data_col, min_row=data_row + 1, max_row=data_row + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showPercent = True
    chart.dLbls.showVal = False
    chart.dLbls.showLeaderLines = True
    if chart.series:
        chart.series[0].data_points = [DataPoint(idx=index) for index in range(len(rows))]
        for index, (_, _, color) in enumerate(rows):
            chart.series[0].data_points[index].graphicalProperties.solidFill = excel_color(color)
    add_chart_with_layout(ws, chart, layout)
    return True


def write_native_chart(ws, title: str, chart_data: dict[str, Any], layout: dict[str, Any], data_row: int, data_col: int) -> bool:
    chart_type = chart_data.get("typ")
    if chart_type == "bar":
        return write_bar_chart(ws, title, chart_data, layout, data_row, data_col)
    if chart_type == "pie":
        return write_pie_chart(ws, title, chart_data, layout, data_row, data_col)
    return write_line_chart(ws, title, chart_data, layout, data_row, data_col)


def write_finstat_charts(ws, graphs: list[dict[str, Any]], start_row: int = 12) -> None:
    if not isinstance(graphs, list):
        return
    data_row = 300
    data_col = 1
    rendered_index = 0
    for index, graph in enumerate(graphs, start=1):
        chart_data = graph.get("chart_data") if isinstance(graph, dict) else None
        if not isinstance(chart_data, dict):
            continue
        title = graph.get("nazov") or chart_data.get("nazov") or f"Graf {index}"
        layout = overview_chart_layout(title, rendered_index)
        if not write_native_chart(ws, str(title), chart_data, layout, data_row, data_col):
            continue
        title_cell = ws[str(layout["title_cell"])]
        title_cell.value = title
        title_cell.alignment = Alignment(wrap_text=True, vertical="top")
        rendered_index += 1
        data_row += max(len(chart_data.get("body") or []) + 4, 8)

def build_template_xlsx(data: dict[str, Any]) -> tuple[BytesIO, str]:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template nenájdený: {TEMPLATE_PATH}")
    workbook = load_workbook(TEMPLATE_PATH)
    write_overview(workbook, data)
    write_orsr(workbook, data)
    write_rpvs(workbook, data)
    write_finstat(workbook, data)
    write_ruz(workbook, data)
    write_crz(workbook, data)
    apply_output_column_widths(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, export_filename(data)
