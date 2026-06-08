from __future__ import annotations

import math
import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image, ImageDraw, ImageFont


TEMPLATE_PATH = Path(__file__).with_name("template.xlsx")
SOURCE_URL_FIELD = "__source_url"

LABELS = {
    "obchodne_meno": "Obchodné meno",
    "sidlo": "Sídlo",
    "den_zapisu": "Deň zápisu",
    "pravna_forma": "Právna forma",
    "vyska_zakladneho_imania": "Výška základného imania",
    "typ_organu": "Typ orgánu",
    "statutarny_organ": "Štatutárny orgán",
    "pravny_predchodca": "Právny predchodca",
    "konanie_menom_spolocnosti": "Konanie menom spoločnosti",
    "dozorna_rada": "Dozorná rada",
    "akcionar": "Akcionár",
    "dalsie_pravne_skutocnosti": "Ďalšie právne skutočnosti",
    "datum_aktualizacie_dat": "Dátum aktualizácie dát",
    "partner_verejneho_sektora": "Partner verejného sektora",
    "opravnena_osoba": "Oprávnená osoba",
    "konecni_uzivatelia_vyhod": "Koneční užívatelia výhod",
    "oznamenie_o_overeni_konecnych_uzivatelov_vyhod": "Oznámenie o overení konečných užívateľov výhod",
    "uctovne_zavierky": "Účtovné závierky",
    "vyrocne_spravy": "Výročné správy",
    "nazov_zmluvy": "Názov zmluvy",
    "cislo_zmluvy": "Číslo zmluvy",
    "dodavatel": "Dodávateľ",
    "odberatel": "Odberateľ",
    "vyhladavanie_podla": "Vyhľadávanie podľa",
    "hladany_dodavatel": "Hľadaný dodávateľ",
}


def export_filename(data: dict[str, Any]) -> str:
    ico = str(data.get("ico") or "export").strip() or "export"
    ico = re.sub(r"[^0-9A-Za-z_-]", "", ico)
    return f"export_{ico}.xlsx"


def label_for(key: Any) -> str:
    text = str(key)
    return LABELS.get(text, text.replace("_", " ").capitalize())


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        if value.startswith("data:image/"):
            return "[obrázok grafu vynechaný]"
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, dict):
        if set(value.keys()) == {"message"}:
            return value.get("message") or ""
        return "\n".join(f"{label_for(k)}: {value_to_text(v)}" for k, v in value.items() if k != SOURCE_URL_FIELD)
    if isinstance(value, list):
        return "\n".join(value_to_text(item) for item in value)
    return str(value)


def fill_value(ws, cell_ref: str, value: Any, hyperlink: str | None = None) -> None:
    cell = ws[cell_ref]
    cell.value = value_to_text(value)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.style = "Hyperlink"


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 7) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def ensure_rows(ws, start_row: int, reserved_rows: int, required_rows: int, max_col: int = 7) -> None:
    if required_rows <= reserved_rows:
        return
    insert_at = start_row + reserved_rows
    extra = required_rows - reserved_rows
    ws.insert_rows(insert_at, extra)
    for row in range(insert_at, insert_at + extra):
        copy_row_style(ws, start_row, row, max_col=max_col)


def clear_row_values(ws, row: int, max_col: int = 7) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def write_pairs(ws, start_row: int, pairs: list[tuple[Any, Any]], reserved_rows: int = 1) -> None:
    rows = pairs or [("", "")]
    ensure_rows(ws, start_row, reserved_rows, len(rows))
    for offset, (key, value) in enumerate(rows):
        row = start_row + offset
        clear_row_values(ws, row)
        ws.cell(row, 1).value = label_for(key) if key else ""
        ws.cell(row, 4).value = value_to_text(value)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")


def write_lines(ws, start_row: int, lines: list[Any], reserved_rows: int = 1) -> None:
    rows = lines or [""]
    ensure_rows(ws, start_row, reserved_rows, len(rows))
    for offset, value in enumerate(rows):
        row = start_row + offset
        clear_row_values(ws, row)
        ws.cell(row, 1).value = value_to_text(value)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")


def dict_pairs(data: dict[str, Any], skip: set[str] | None = None) -> list[tuple[str, Any]]:
    skip = skip or set()
    return [(key, value) for key, value in data.items() if key not in skip and key != SOURCE_URL_FIELD]


def row_text(row: Any) -> str:
    if isinstance(row, list):
        return "\n".join(value_to_text(item) for item in row)
    return value_to_text(row)


def write_record_lines(ws, start_row: int, records: list[Any], reserved_rows: int = 1) -> None:
    write_lines(ws, start_row, [row_text(record) for record in records], reserved_rows=reserved_rows)


def write_table(
    ws,
    start_row: int,
    records: list[dict[str, Any]],
    columns: list[tuple[str, str]],
    reserved_rows: int,
    hyperlink_getter=None,
) -> None:
    rows = records or []
    ensure_rows(ws, start_row, reserved_rows, max(len(rows), 1))
    if not rows:
        clear_row_values(ws, start_row)
        ws.cell(start_row, 1).value = ""
        return
    for offset, record in enumerate(rows):
        row_num = start_row + offset
        clear_row_values(ws, row_num)
        for col_index, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row_num, col_index)
            cell.value = value_to_text(record.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            hyperlink = hyperlink_getter(record, key) if hyperlink_getter else None
            if hyperlink:
                cell.hyperlink = hyperlink
                cell.style = "Hyperlink"


def find_row_with_value(ws, value: str) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == value:
                return cell.row
    return None


def clean_labeled_value(value: Any, labels: tuple[str, ...]) -> str:
    text = value_to_text(value).strip()
    for label in labels:
        text = re.sub(rf"^\s*{re.escape(label)}\s*:?\s*", "", text, flags=re.IGNORECASE)
    return text.strip()


def clean_od_value(value: Any) -> str:
    text = clean_labeled_value(value, ("Od", "od"))
    match = re.search(r"\(?\s*od\s*:?\s*([^)]*)\)?", text, flags=re.IGNORECASE)
    if match:
        text = match.group(1)
    return text.strip().strip("()")


def parse_orsr_stock_row(row: Any) -> dict[str, str]:
    values = row if isinstance(row, list) else [row]
    parsed = {"pocet": "", "druh": "", "podoba": "", "forma": "", "menovita_hodnota": "", "od": ""}
    unlabeled = []

    for item in values:
        text = value_to_text(item).strip()
        if not text:
            continue
        lowered = text.casefold()
        if lowered.startswith("počet") or lowered.startswith("pocet"):
            parsed["pocet"] = clean_labeled_value(text, ("Počet", "Pocet"))
        elif lowered.startswith("druh"):
            parsed["druh"] = clean_labeled_value(text, ("Druh",))
        elif lowered.startswith("podoba"):
            parsed["podoba"] = clean_labeled_value(text, ("Podoba",))
        elif lowered.startswith("forma"):
            parsed["forma"] = clean_labeled_value(text, ("Forma",))
        elif lowered.startswith("menovitá hodnota") or lowered.startswith("menovita hodnota"):
            parsed["menovita_hodnota"] = clean_labeled_value(text, ("Menovitá hodnota", "Menovita hodnota"))
        elif re.search(r"\(?\s*od\s*:", lowered) or lowered.startswith("od"):
            parsed["od"] = clean_od_value(text)
        else:
            unlabeled.append(text)

    ordered_keys = ["pocet", "druh", "podoba", "forma", "menovita_hodnota", "od"]
    for key, value in zip([key for key in ordered_keys if not parsed[key]], unlabeled):
        parsed[key] = clean_od_value(value) if key == "od" else value

    return parsed


def write_orsr_stocks_table(ws, start_row: int, stocks: list[Any], reserved_rows: int = 1) -> None:
    rows = [parse_orsr_stock_row(stock) for stock in (stocks or [])]
    ensure_rows(ws, start_row, reserved_rows, max(len(rows), 1))
    if not rows:
        clear_row_values(ws, start_row)
        return
    columns = ["pocet", "druh", "podoba", "forma", "menovita_hodnota", "od"]
    for offset, record in enumerate(rows):
        row_num = start_row + offset
        clear_row_values(ws, row_num)
        for col_index, key in enumerate(columns, start=1):
            ws.cell(row_num, col_index).value = record.get(key, "")
            ws.cell(row_num, col_index).alignment = Alignment(wrap_text=True, vertical="top")


def get_ci(data: dict[str, Any], *keys: str) -> Any:
    if not isinstance(data, dict):
        return ""
    normalized = {str(key).casefold(): value for key, value in data.items()}
    for key in keys:
        value = normalized.get(key.casefold())
        if value not in (None, ""):
            return value
    return ""


def first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def first_document_url(record: dict[str, Any]) -> str:
    documents = record.get("dokumenty") if isinstance(record, dict) else None
    if not isinstance(documents, list):
        return ""
    for document in documents:
        if isinstance(document, dict) and document.get("url"):
            return str(document["url"])
    return ""


def ruz_table_hyperlink(record: dict[str, Any], key: str) -> str:
    if key == "Obdobie":
        return first_document_url(record)
    return ""


def crz_table_hyperlink(record: dict[str, Any], key: str) -> str:
    if key == "nazov_zmluvy" and isinstance(record, dict):
        return str(record.get("link") or "")
    return ""


def write_overview(wb, data: dict[str, Any]) -> None:
    ws = wb["Prehľad"]
    orsr = data.get("orsr", {}) if isinstance(data.get("orsr"), dict) else {}
    finstat_basic = data.get("finstat", {}).get("zakladne_udaje", {}) if isinstance(data.get("finstat"), dict) else {}
    fill_value(ws, "C5", data.get("ico", ""))
    fill_value(ws, "C6", first_value(orsr.get("obchodne_meno"), finstat_basic.get("Obchodné meno"), finstat_basic.get("Názov")))
    fill_value(ws, "C7", first_value(orsr.get("sidlo"), finstat_basic.get("Sídlo"), finstat_basic.get("Adresa")))
    fill_value(ws, "C8", first_value(orsr.get("den_zapisu"), finstat_basic.get("Dátum vzniku")))
    ws["A12"].value = None
    write_finstat_charts(ws, data.get("finstat", {}).get("grafy", []), start_row=12)


def write_orsr(wb, data: dict[str, Any]) -> None:
    orsr = data.get("orsr", {}) if isinstance(data.get("orsr"), dict) else {}
    ws = wb["ORSR"]
    fixed = {
        "D2": orsr.get("obchodne_meno"),
        "D3": orsr.get("sidlo"),
        "D4": orsr.get("den_zapisu"),
        "D5": orsr.get("pravna_forma"),
        "D6": orsr.get("vyska_zakladneho_imania"),
        "D7": orsr.get("typ_organu"),
        "A23": orsr.get("datum_aktualizacie_dat"),
    }
    for cell, value in fixed.items():
        fill_value(ws, cell, value)
    write_record_lines(ws, 9, orsr.get("statutarny_organ", []), reserved_rows=7)
    write_record_lines(ws, 17, orsr.get("konanie_menom_spolocnosti", []), reserved_rows=1)
    write_record_lines(ws, 19, orsr.get("dozorna_rada", []), reserved_rows=3)
    fill_value(ws, "A25", orsr.get(SOURCE_URL_FIELD, ""), hyperlink=orsr.get(SOURCE_URL_FIELD, ""))

    ws = wb["ORSR - Akcie"]
    write_record_lines(ws, 3, orsr.get("akcionar", []), reserved_rows=1)
    stock_title_row = find_row_with_value(ws, "Akcie") or 4
    stock_header_row = stock_title_row + 1
    headers = ["Počet", "Druh", "Podoba", "Forma", "Menovitá hodnota", "Od"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(stock_header_row, col_index).value = header
    write_orsr_stocks_table(ws, stock_header_row + 1, orsr.get("akcie", []), reserved_rows=1)

    ws = wb["ORSR - Podnikanie"]
    write_lines(ws, 2, orsr.get("predmet_podnikania", []), reserved_rows=1)

    ws = wb["ORSR - Právne skutočnosti"]
    lines = []
    lines.extend(row_text(row) for row in orsr.get("pravny_predchodca", []))
    lines.extend(row_text(row) for row in orsr.get("dalsie_pravne_skutocnosti", []))
    write_lines(ws, 2, lines, reserved_rows=1)


def write_rpvs(wb, data: dict[str, Any]) -> None:
    rpvs = data.get("rpvs", {}) if isinstance(data.get("rpvs"), dict) else {}
    ws = wb["RPVS"]
    partner = rpvs.get("partner_verejneho_sektora", {}) if isinstance(rpvs.get("partner_verejneho_sektora"), dict) else {}
    mapping = {
        "D2": get_ci(partner, "Obchodné meno", "Názov", "Meno / názov"),
        "D3": get_ci(partner, "IČO"),
        "D4": get_ci(partner, "Právna forma"),
        "D5": get_ci(partner, "Adresa", "Sídlo"),
        "D6": get_ci(partner, "Dátum zápisu"),
        "D7": get_ci(partner, "Dátum výmazu"),
        "D8": get_ci(partner, "Číslo vložky"),
    }
    for cell, value in mapping.items():
        fill_value(ws, cell, value)
    opravnena_heading_row = find_row_with_value(ws, "Oprávnená osoba") or 9
    opravnena_pairs = dict_pairs(rpvs.get("opravnena_osoba", {}))
    opravnena_start_row = opravnena_heading_row + 1
    write_pairs(ws, opravnena_start_row, opravnena_pairs, reserved_rows=1)

    spacer_row = opravnena_start_row + max(len(opravnena_pairs), 1)
    ws.insert_rows(spacer_row, 1)
    clear_row_values(ws, spacer_row)

    kuvy_heading_row = find_row_with_value(ws, "Koneční užívatelia výhod") or 11
    kuvy_start_row = kuvy_heading_row + 2
    oznamenie_heading_row = find_row_with_value(ws, "Oznámenie o overení konečných užívateľov výhod") or 19
    kuvy_reserved_rows = max(1, oznamenie_heading_row - kuvy_start_row)
    write_table(ws, kuvy_start_row, rpvs.get("konecni_uzivatelia_vyhod", []), [
        ("Meno a priezvisko", "Meno a priezvisko"),
        ("Dátum narodenia", "Dátum narodenia"),
        ("Štátna príslušnosť", "Štátna príslušnosť"),
        ("Adresa", "Adresa"),
        ("Verejný funkcionár", "Verejný funkcionár"),
    ], reserved_rows=kuvy_reserved_rows)

    oznamenie_heading_row = find_row_with_value(ws, "Oznámenie o overení konečných užívateľov výhod") or oznamenie_heading_row
    zdroj_heading_row = find_row_with_value(ws, "Zdroj") or 25
    oznamenie_reserved_rows = max(1, zdroj_heading_row - (oznamenie_heading_row + 1))
    write_pairs(
        ws,
        oznamenie_heading_row + 1,
        dict_pairs(rpvs.get("oznamenie_o_overeni_konecnych_uzivatelov_vyhod", {})),
        reserved_rows=oznamenie_reserved_rows,
    )

    zdroj_heading_row = find_row_with_value(ws, "Zdroj") or zdroj_heading_row
    fill_value(ws, f"A{zdroj_heading_row + 1}", rpvs.get(SOURCE_URL_FIELD, ""), hyperlink=rpvs.get(SOURCE_URL_FIELD, ""))


def write_finstat(wb, data: dict[str, Any]) -> None:
    finstat = data.get("finstat", {}) if isinstance(data.get("finstat"), dict) else {}
    ws = wb["Finstat"]
    basic = finstat.get("zakladne_udaje", {}) if isinstance(finstat.get("zakladne_udaje"), dict) else {}
    financial = finstat.get("financne_ukazovatele", {}) if isinstance(finstat.get("financne_ukazovatele"), dict) else {}
    important = finstat.get("dolezite_ukazovatele", {}) if isinstance(finstat.get("dolezite_ukazovatele"), dict) else {}
    for row in range(3, 11):
        fill_value(ws, f"D{row}", get_ci(basic, str(ws[f"A{row}"].value or "")))
    for row in range(12, 19):
        fill_value(ws, f"D{row}", get_ci(financial, str(ws[f"A{row}"].value or "")))
    for row in range(20, 29):
        label = str(ws[f"A{row}"].value or "").replace(";", "")
        fill_value(ws, f"D{row}", get_ci(important, label))
    fill_value(ws, "A31", finstat.get(SOURCE_URL_FIELD, ""), hyperlink=finstat.get(SOURCE_URL_FIELD, ""))


def write_ruz(wb, data: dict[str, Any]) -> None:
    ruz = data.get("ruz", {}) if isinstance(data.get("ruz"), dict) else {}
    ws = wb["RÚZ"]
    for row in range(2, 14):
        fill_value(ws, f"D{row}", get_ci(ruz, str(ws[f"A{row}"].value or "")))
    statements = ruz.get("uctovne_zavierky") or ruz.get("zaznamy") or []
    annual_reports = ruz.get("vyrocne_spravy") or []
    statement_type = "Typ závierky"
    annual_type = "Typ výročnej správy"
    write_table(ws, 16, statements, [
        ("Obdobie", "Obdobie"), (statement_type, statement_type), ("Predložená dňa", "Predložená dňa"),
        ("Zostavená dňa", "Zostavená dňa"), ("Schválená dňa", "Schválená dňa"), ("SA uložená dňa", "SA uložená dňa"),
    ], reserved_rows=13, hyperlink_getter=ruz_table_hyperlink)
    write_table(ws, 31, annual_reports, [
        ("Obdobie", "Obdobie"), (annual_type, annual_type), ("Predložená dňa", "Predložená dňa"),
        ("Zostavená dňa", "Zostavená dňa"), ("Schválená dňa", "Schválená dňa"), ("SA uložená dňa", "SA uložená dňa"),
    ], reserved_rows=16, hyperlink_getter=ruz_table_hyperlink)
    fill_value(ws, "A48", ruz.get(SOURCE_URL_FIELD, ""), hyperlink=ruz.get(SOURCE_URL_FIELD, ""))


def write_crz(wb, data: dict[str, Any]) -> None:
    crz = data.get("crz", {}) if isinstance(data.get("crz"), dict) else {}
    ws = wb["CRZ"]
    write_table(ws, 3, crz.get("zmluvy", []), [
        ("datum", "Dátum"), ("nazov_zmluvy", "Zmluva"), ("cena", "Cena"),
        ("dodavatel", "Dodávateľ"), ("odberatel", "Odberateľ"),
    ], reserved_rows=11, hyperlink_getter=crz_table_hyperlink)
    fill_value(ws, "A15", crz.get(SOURCE_URL_FIELD, ""), hyperlink=crz.get(SOURCE_URL_FIELD, ""))


def numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    match = re.search(r"-?\d[\d ]*(?:[,.]\d+)?", value)
    if not match:
        return None
    number = float(match.group(0).replace(" ", "").replace(",", "."))
    lowered = value.lower()
    if "mld" in lowered:
        number *= 1_000_000_000
    elif "mil" in lowered:
        number *= 1_000_000
    elif "tis" in lowered:
        number *= 1_000
    return number


def chart_font(size: int = 12):
    try:
        return ImageFont.truetype("DejaVuSans.ttf", size)
    except Exception:
        return ImageFont.load_default()


def rgb(color: Any, fallback: str = "#52606d") -> tuple[int, int, int]:
    value = str(color or fallback).strip()
    if value.startswith("rgb"):
        nums = [int(float(num)) for num in re.findall(r"\d+(?:\.\d+)?", value)[:3]]
        if len(nums) == 3:
            return tuple(nums)
    value = value.lstrip("#")
    if len(value) == 3:
        value = "".join(ch * 2 for ch in value)
    if len(value) == 6:
        try:
            return tuple(int(value[index:index + 2], 16) for index in (0, 2, 4))
        except ValueError:
            pass
    return rgb(fallback, "#52606d")


def draw_wrapped_text(draw, xy: tuple[int, int], text: str, font, fill=(50, 63, 75), max_width: int = 220, line_height: int = 15) -> int:
    words = str(text or "").split()
    if not words:
        return 0
    x, y = xy
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    for offset, line in enumerate(lines):
        draw.text((x, y + offset * line_height), line, font=font, fill=fill)
    return len(lines) * line_height


def png_bytes(image: Image.Image) -> BytesIO:
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def render_line_chart_image(title: str, chart_data: dict[str, Any]) -> BytesIO | None:
    points = chart_data.get("body") or []
    if len(points) < 2:
        return None
    width, height = 520, 260
    padding = {"top": 28, "right": 30, "bottom": 44, "left": 34}
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font = chart_font(11)
    value_font = chart_font(11)
    xs = [point.get("x_svg") if isinstance(point.get("x_svg"), (int, float)) else index for index, point in enumerate(points)]
    ys = [point.get("y_svg") if isinstance(point.get("y_svg"), (int, float)) else 0 for point in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    plot_width = width - padding["left"] - padding["right"]
    plot_height = height - padding["top"] - padding["bottom"]

    def sx(value):
        return padding["left"] + ((value - min_x) / max(max_x - min_x, 1)) * plot_width

    def sy(value):
        return padding["top"] + ((value - min_y) / max(max_y - min_y, 1)) * plot_height

    for ratio in (0, 0.33, 0.66, 1):
        y = padding["top"] + ratio * plot_height
        for x in range(padding["left"], width - padding["right"], 12):
            draw.line((x, y, min(x + 6, width - padding["right"]), y), fill=rgb("#e6e5e5"), width=1)

    chart_points = [(sx(xs[index]), sy(ys[index]), point) for index, point in enumerate(points)]
    draw.line([(x, y) for x, y, _ in chart_points], fill=rgb("#0f766e"), width=3, joint="curve")
    for x, y, point in chart_points:
        draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill="white", outline=rgb("#0f766e"), width=2)
        if point.get("value_label"):
            label = str(point.get("value_label"))
            bbox = draw.textbbox((0, 0), label, font=value_font)
            draw.text((x - (bbox[2] - bbox[0]) / 2, max(2, y - 24)), label, font=value_font, fill=rgb("#1f2933"))
        year = str(point.get("rok") or "")
        bbox = draw.textbbox((0, 0), year, font=font)
        draw.text((x - (bbox[2] - bbox[0]) / 2, height - 28), year, font=font, fill=rgb("#52606d"))
    draw.rectangle((0, 0, width - 1, height - 1), outline=rgb("#edf1f5"))
    return png_bytes(image)


def render_bar_chart_image(title: str, chart_data: dict[str, Any]) -> BytesIO | None:
    rows = chart_data.get("body") or []
    legend = chart_data.get("legend") or []
    if not rows or not legend:
        return None
    width, chart_height, legend_height = 760, 340, 110
    height = chart_height + legend_height
    padding = {"top": 34, "right": 28, "bottom": 48, "left": 34}
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font = chart_font(11)
    legend_font = chart_font(12)
    plot_width = width - padding["left"] - padding["right"]
    plot_height = chart_height - padding["top"] - padding["bottom"]
    max_stack = max([float(row.get("stack_height_svg") or 0) for row in rows] + [1])
    bar_width = min(76, plot_width / max(len(rows), 1) * 0.6)
    gap = plot_width / (len(rows) - 1) if len(rows) > 1 else 0

    for ratio in (0, 0.33, 0.66, 1):
        y = padding["top"] + ratio * plot_height
        for x in range(padding["left"], width - padding["right"], 12):
            draw.line((x, y, min(x + 6, width - padding["right"]), y), fill=rgb("#e6e5e5"), width=1)

    for row_index, row in enumerate(rows):
        center_x = padding["left"] + row_index * gap if len(rows) > 1 else padding["left"] + plot_width / 2
        current_bottom = padding["top"] + plot_height
        segments = [segment for segment in row.get("segments", []) if float(segment.get("height_svg") or 0) > 0]
        for segment in reversed(segments):
            segment_height = max(1, (float(segment.get("height_svg") or 0) / max_stack) * plot_height)
            y = current_bottom - segment_height
            draw.rectangle((center_x - bar_width / 2, y, center_x + bar_width / 2, current_bottom), fill=rgb(segment.get("color")))
            current_bottom = y
        if row.get("total_label"):
            label = str(row.get("total_label"))
            bbox = draw.textbbox((0, 0), label, font=font)
            draw.text((center_x - (bbox[2] - bbox[0]) / 2, max(2, current_bottom - 18)), label, font=font, fill=rgb("#1f2933"))
        year = str(row.get("rok") or "")
        bbox = draw.textbbox((0, 0), year, font=font)
        draw.text((center_x - (bbox[2] - bbox[0]) / 2, chart_height - 28), year, font=font, fill=rgb("#52606d"))

    legend_x, legend_y = 12, chart_height + 8
    column_width = 245
    for index, item in enumerate(legend):
        col = index % 3
        row = index // 3
        x = legend_x + col * column_width
        y = legend_y + row * 24
        draw.ellipse((x, y + 4, x + 10, y + 14), fill=rgb(item.get("color")))
        draw_wrapped_text(draw, (x + 18, y), str(item.get("label") or ""), legend_font, max_width=column_width - 24, line_height=14)
    draw.rectangle((0, 0, width - 1, height - 1), outline=rgb("#edf1f5"))
    return png_bytes(image)


def render_pie_chart_image(title: str, chart_data: dict[str, Any]) -> BytesIO | None:
    slices = chart_data.get("body") or []
    rows = []
    for item in slices:
        value = numeric(item.get("value")) or numeric(item.get("value_label"))
        if value is not None and value > 0:
            rows.append((item, value))
    if not rows:
        return None
    width, height = 560, 270
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font = chart_font(12)
    value_font = chart_font(11)
    total = sum(value for _, value in rows) or 1
    box = (32, 42, 222, 232)
    start = -90
    for item, value in rows:
        extent = 360 * value / total
        draw.pieslice(box, start=start, end=start + extent, fill=rgb(item.get("color")))
        start += extent
    draw.ellipse(box, outline="white", width=2)
    inner = (88, 98, 166, 176)
    draw.ellipse(inner, fill="white", outline="white")
    legend_x, legend_y = 250, 22
    for index, (item, _) in enumerate(rows):
        y = legend_y + index * 25
        draw.ellipse((legend_x, y + 5, legend_x + 10, y + 15), fill=rgb(item.get("color")))
        draw_wrapped_text(draw, (legend_x + 18, y), str(item.get("label") or ""), font, max_width=210, line_height=14)
        draw.text((width - 86, y), str(item.get("value_label") or ""), font=value_font, fill=rgb("#52606d"))
    draw.rectangle((0, 0, width - 1, height - 1), outline=rgb("#edf1f5"))
    return png_bytes(image)


def render_chart_image(title: str, chart_data: dict[str, Any]) -> BytesIO | None:
    chart_type = chart_data.get("typ")
    if chart_type == "bar":
        return render_bar_chart_image(title, chart_data)
    if chart_type == "pie":
        return render_pie_chart_image(title, chart_data)
    return render_line_chart_image(title, chart_data)


def write_finstat_charts(ws, graphs: list[dict[str, Any]], start_row: int = 12) -> None:
    if not isinstance(graphs, list):
        return
    chart_row = start_row
    for index, graph in enumerate(graphs, start=1):
        chart_data = graph.get("chart_data") if isinstance(graph, dict) else None
        if not isinstance(chart_data, dict):
            continue
        title = graph.get("nazov") or chart_data.get("nazov") or f"Graf {index}"
        image_data = render_chart_image(title, chart_data)
        if not image_data:
            continue
        ws.cell(chart_row, 1).value = title
        ws.cell(chart_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        xl_image = XLImage(image_data)
        xl_image.anchor = f"A{chart_row + 1}"
        ws.add_image(xl_image)
        row_span = 24 if chart_data.get("typ") == "bar" else 18
        for row in range(chart_row + 1, chart_row + row_span):
            ws.row_dimensions[row].height = 18
        chart_row += row_span


def build_template_xlsx(data: dict[str, Any]) -> tuple[BytesIO, str]:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template nenájdený: {TEMPLATE_PATH}")
    workbook = load_workbook(TEMPLATE_PATH)
    write_overview(workbook, data)
    write_orsr(workbook, data)
    write_rpvs(workbook, data)
    write_finstat(workbook, data)
    write_ruz(workbook, data)
    write_crz(workbook, data)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, export_filename(data)
