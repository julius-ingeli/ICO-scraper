import asyncio
import os
import re
from io import BytesIO
from typing import Any

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from starlette.concurrency import run_in_threadpool

from main import AVAILABLE_SOURCES, scrape_subject


app = FastAPI(title="ICO Scraper", version="1.0.0")
_scrape_limit = asyncio.Semaphore(int(os.getenv("MAX_CONCURRENT_SCRAPES", "1")))


INDEX_HTML = """<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>ICO Scraper</title>
  <style>
    :root {
      color-scheme: light;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f4f6f8;
      color: #1f2933;
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      display: flex;
      justify-content: center;
      padding: 32px;
    }

    main {
      width: min(1120px, 100%);
      display: grid;
      grid-template-rows: auto auto 1fr;
      gap: 18px;
    }

    .topbar {
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 24px;
    }

    h1 {
      margin: 0 0 6px;
      font-size: 32px;
      line-height: 1.1;
      font-weight: 700;
    }

    p {
      margin: 0;
      color: #52606d;
      font-size: 15px;
    }

    form {
      display: flex;
      gap: 10px;
      align-items: center;
      flex-wrap: wrap;
      justify-content: flex-end;
    }

    .source-controls {
      width: 100%;
      display: flex;
      justify-content: flex-end;
      gap: 8px;
      flex-wrap: wrap;
    }

    .source-toggle {
      height: 34px;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      border: 1px solid #cbd2d9;
      border-radius: 6px;
      padding: 0 10px;
      background: white;
      color: #323f4b;
      font-size: 13px;
      font-weight: 700;
      cursor: pointer;
      user-select: none;
    }

    .source-toggle input {
      width: 16px;
      height: 16px;
      margin: 0;
      accent-color: #0f766e;
    }

    label {
      font-size: 14px;
      font-weight: 650;
      color: #323f4b;
    }

    input {
      width: 180px;
      height: 42px;
      border: 1px solid #cbd2d9;
      border-radius: 6px;
      padding: 0 12px;
      font-size: 16px;
      background: white;
      color: #1f2933;
    }

    button {
      height: 42px;
      border: 0;
      border-radius: 6px;
      padding: 0 16px;
      font-size: 15px;
      font-weight: 700;
      background: #0f766e;
      color: white;
      cursor: pointer;
    }

    button:disabled {
      cursor: progress;
      opacity: 0.68;
    }

    .status {
      min-height: 22px;
      font-size: 14px;
      color: #52606d;
    }

    .status.error { color: #b42318; }

    .tabs {
      display: none;
      gap: 8px;
      align-items: center;
      overflow-x: auto;
      padding-bottom: 2px;
      border-bottom: 1px solid #d9e2ec;
    }

    .tabs.visible { display: flex; }

    .tab-button {
      flex: 0 0 auto;
      height: 38px;
      border: 1px solid #cbd2d9;
      border-bottom: 0;
      border-radius: 6px 6px 0 0;
      padding: 0 14px;
      background: #e4e7eb;
      color: #323f4b;
      font-size: 14px;
      font-weight: 700;
      cursor: pointer;
    }

    .tab-button.active {
      background: white;
      color: #0f766e;
      border-color: #9fb3c8;
    }

    .results {
      min-height: 420px;
      border: 1px solid #d9e2ec;
      border-radius: 8px;
      background: white;
      overflow: hidden;
    }

    .empty-state {
      min-height: 420px;
      display: grid;
      place-items: center;
      padding: 28px;
      color: #52606d;
      text-align: center;
    }

    .panel {
      display: none;
      padding: 20px;
    }

    .panel.active { display: block; }

    .panel-title {
      margin: 0 0 16px;
      font-size: 22px;
      line-height: 1.25;
      color: #1f2933;
    }

    .summary-layout {
      display: grid;
      gap: 22px;
    }

    .finstat-graphs {
      display: grid;
      gap: 14px;
    }

    .finstat-graphs-title {
      margin: 0;
      font-size: 18px;
      line-height: 1.3;
      color: #1f2933;
    }

    .finstat-graph-list {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 14px;
    }

    .finstat-graph {
      border: 1px solid #d9e2ec;
      border-radius: 8px;
      padding: 12px;
      background: #f8fafc;
    }

    .finstat-graph-title {
      margin: 0 0 10px;
      font-size: 14px;
      line-height: 1.35;
      color: #323f4b;
    }

    .finstat-graph img {
      display: block;
      width: 100%;
      height: auto;
      border: 1px solid #edf1f5;
      border-radius: 6px;
      background: white;
    }

    .section-subtitle {
      margin: 18px 0 10px;
      font-size: 18px;
      line-height: 1.3;
      color: #1f2933;
    }

    .notice {
      padding: 14px 16px;
      border: 1px solid #f0c36d;
      border-radius: 6px;
      background: #fff8e6;
      color: #7c4a03;
      font-size: 14px;
      line-height: 1.45;
    }

    .field-grid {
      display: grid;
      grid-template-columns: minmax(180px, 280px) minmax(0, 1fr);
      gap: 0;
      border-top: 1px solid #edf1f5;
    }

    .field-label,
    .field-value {
      padding: 12px 10px;
      border-bottom: 1px solid #edf1f5;
      font-size: 14px;
      line-height: 1.45;
    }

    .field-label {
      color: #52606d;
      font-weight: 700;
    }

    .field-value {
      color: #1f2933;
      overflow-wrap: anywhere;
    }

    .nested {
      margin: 4px 0 0;
      padding-left: 14px;
      border-left: 3px solid #d9e2ec;
    }

    .list {
      display: grid;
      gap: 10px;
    }

    .list-item {
      padding: 12px;
      border: 1px solid #d9e2ec;
      border-radius: 6px;
      background: #f8fafc;
    }

    .data-table-wrap {
      overflow-x: auto;
      border: 1px solid #d9e2ec;
      border-radius: 8px;
      background: white;
    }

    .data-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
      line-height: 1.4;
    }

    .data-table th,
    .data-table td {
      padding: 11px 12px;
      border-bottom: 1px solid #edf1f5;
      text-align: left;
      vertical-align: top;
    }

    .data-table th {
      background: #f8fafc;
      color: #52606d;
      font-weight: 700;
      white-space: nowrap;
    }

    .period-button {
      border: 0;
      background: transparent;
      color: #0f766e;
      padding: 0;
      height: auto;
      font: inherit;
      font-weight: 700;
      text-decoration: underline;
      cursor: pointer;
    }

    .documents-row { display: none; }
    .documents-row.visible { display: table-row; }

    .documents-cell {
      background: #f8fafc;
      padding: 14px 16px !important;
    }

    .documents-title {
      margin: 0 0 10px;
      color: #323f4b;
      font-size: 14px;
      font-weight: 700;
    }

    .document-link {
      color: #0f766e;
      font-weight: 700;
      text-decoration: underline;
    }

    .export-panel {
      min-height: 320px;
      display: grid;
      align-content: start;
      gap: 14px;
      padding: 18px;
      border: 1px solid #d9e2ec;
      border-radius: 8px;
      background: #f8fafc;
    }

    .export-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }

    .export-button {
      height: 42px;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      border: 1px solid #0f766e;
      border-radius: 6px;
      padding: 0 14px;
      background: #0f766e;
      color: white;
      font-size: 14px;
      font-weight: 700;
      cursor: pointer;
    }

    .export-button.secondary {
      background: white;
      color: #0f766e;
    }

    .export-button svg {
      width: 18px;
      height: 18px;
      stroke: currentColor;
      stroke-width: 2;
      fill: none;
      stroke-linecap: round;
      stroke-linejoin: round;
    }

    @media (max-width: 720px) {
      body { padding: 18px; }
      .topbar { align-items: stretch; flex-direction: column; }
      h1 { font-size: 26px; }
      form { align-items: stretch; justify-content: stretch; }
      label { width: 100%; }
      input, button { width: 100%; }
      .source-controls { justify-content: stretch; }
      .source-toggle { flex: 1 1 calc(50% - 8px); }
      .source-toggle input { width: 16px; }
      .field-grid { grid-template-columns: 1fr; }
      .field-label { padding-bottom: 2px; border-bottom: 0; }
      .field-value { padding-top: 2px; }
    }
  </style>
</head>
<body>
  <main>
    <section class="topbar">
      <div>
        <h1>ICO Scraper</h1>
        <p>Zadajte IČO a aplikácia rozdelí výsledok podľa zdroja.</p>
      </div>
      <form id="scrape-form">
        <label for="ico">IČO</label>
        <input id="ico" name="ico" inputmode="numeric" autocomplete="off" pattern="[0-9]{8}" maxlength="8" placeholder="36785512" required />
        <button id="submit" type="submit">Vyhľadať</button>
        <div class="source-controls" aria-label="Zdroje informácií">
          <label class="source-toggle"><input type="checkbox" name="source" value="orsr" checked /> ORSR</label>
          <label class="source-toggle"><input type="checkbox" name="source" value="rpvs" checked /> RPVS</label>
          <label class="source-toggle"><input type="checkbox" name="source" value="finstat" checked /> FinStat</label>
          <label class="source-toggle"><input type="checkbox" name="source" value="ruz" checked /> RÚZ</label>
          <label class="source-toggle"><input type="checkbox" name="source" value="crz" checked /> CRZ</label>
        </div>
      </form>
    </section>
    <div class="status" id="status"></div>
    <nav class="tabs" id="tabs" aria-label="Kategórie výsledku"></nav>
    <section class="results" id="results">
      <div class="empty-state">Výsledky sa zobrazia po vyhľadaní IČO.</div>
    </section>
  </main>

  <script>
    const form = document.querySelector('#scrape-form');
    const input = document.querySelector('#ico');
    const button = document.querySelector('#submit');
    const statusEl = document.querySelector('#status');
    const tabs = document.querySelector('#tabs');
    const results = document.querySelector('#results');
    const sourceInputs = Array.from(document.querySelectorAll('input[name="source"]'));
    let latestRawJson = '';
    let latestData = null;

    const downloadIcon = `
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path>
        <path d="M7 10l5 5 5-5"></path>
        <path d="M12 15V3"></path>
      </svg>
    `;

    const copyIcon = `
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <rect width="14" height="14" x="8" y="8" rx="2" ry="2"></rect>
        <path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"></path>
      </svg>
    `;

    const categoryLabels = {
      ico: 'Všeob. Info',
      orsr: 'ORSR',
      rpvs: 'RPVS',
      finstat: 'FinStat',
      ruz: 'RÚZ',
      crz: 'CRZ',
      raw: 'Export dát'
    };

    function formatKey(key) {
      return categoryLabels[key] || key
        .replaceAll('_', ' ')
        .replace(/\\b\\w/g, letter => letter.toUpperCase());
    }

    function selectedSources() {
      return sourceInputs
        .filter(input => input.checked)
        .map(input => input.value);
    }

    function clearResults(message = 'Výsledky sa zobrazia po vyhľadaní IČO.') {
      latestData = null;
      tabs.className = 'tabs';
      tabs.innerHTML = '';
      results.innerHTML = `<div class="empty-state">${message}</div>`;
    }

    function renderValue(value) {
      if (value === null || value === undefined || value === '') {
        return '<span>-</span>';
      }

      if (typeof value === 'object' && !Array.isArray(value) && Object.keys(value).length === 1 && value.message) {
        return `<div class="notice">${renderValue(value.message)}</div>`;
      }

      if (typeof value === 'object' && !Array.isArray(value) && value.image) {
        return `
          <article class="finstat-graph">
            <h4 class="finstat-graph-title">${renderValue(value.nazov || 'Graf')}</h4>
            <img src="${escapeAttribute(value.image)}" alt="${escapeAttribute(value.nazov || 'FinStat graf')}" loading="lazy" />
          </article>
        `;
      }

      if (Array.isArray(value)) {
        if (value.length === 0) return '<span>-</span>';
        return `<div class="list">${value.map(item => `
          <div class="list-item">${renderValue(item)}</div>
        `).join('')}</div>`;
      }

      if (typeof value === 'object') {
        const entries = Object.entries(value);
        if (entries.length === 0) return '<span>-</span>';
        return `<div class="field-grid nested">${entries.map(([key, child]) => `
          <div class="field-label">${formatKey(key)}</div>
          <div class="field-value">${renderValue(child)}</div>
        `).join('')}</div>`;
      }

      return String(value)
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;');
    }

    function escapeAttribute(value) {
      return String(value)
        .replaceAll('&', '&amp;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#39;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;');
    }


    function firstValue(...values) {
      for (const value of values) {
        if (value !== null && value !== undefined && String(value).trim() !== '') {
          return value;
        }
      }
      return '-';
    }

    function renderSummary(data) {
      const orsr = data.orsr || {};
      const finstat = data.finstat || {};
      const finstatBasic = finstat.zakladne_udaje || {};
      const companyName = firstValue(orsr.obchodne_meno, finstatBasic['Obchodné meno'], finstatBasic.Názov);
      const address = firstValue(orsr.sidlo, finstatBasic['Sídlo'], finstatBasic.Adresa);
      const foundingDay = firstValue(orsr.den_zapisu, finstatBasic['Dátum vzniku']);

      return `
        <section class="summary-layout">
          <div class="field-grid">
            <div class="field-label">IČO</div>
            <div class="field-value">${renderValue(data.ico)}</div>
            <div class="field-label">Názov spoločnosti</div>
            <div class="field-value">${renderValue(companyName)}</div>
            <div class="field-label">Adresa</div>
            <div class="field-value">${renderValue(address)}</div>
            <div class="field-label">Deň vzniku</div>
            <div class="field-value">${renderValue(foundingDay)}</div>
          </div>
          ${'finstat' in data ? renderFinstatGraphs(finstat.grafy || []) : ''}
        </section>
      `;
    }

    function renderFinstatGraphs(graphs) {
      if (!Array.isArray(graphs) || graphs.length === 0) {
        return '<div class="notice">FinStat grafy sa nepodarilo načítať.</div>';
      }

      return `
        <section class="finstat-graphs" aria-label="FinStat grafy">
          <h3 class="finstat-graphs-title">FinStat grafy</h3>
          <div class="finstat-graph-list">
            ${graphs.map((graph, index) => `
              <article class="finstat-graph">
                <h4 class="finstat-graph-title">${renderValue(graph.nazov || `Graf ${index + 1}`)}</h4>
                <img src="${escapeAttribute(graph.image)}" alt="${escapeAttribute(graph.nazov || `FinStat graf ${index + 1}`)}" loading="lazy" />
              </article>
            `).join('')}
          </div>
        </section>
      `;
    }

    async function copyText(text) {
      if (navigator.clipboard && window.isSecureContext) {
        await navigator.clipboard.writeText(text);
        return;
      }

      const textarea = document.createElement('textarea');
      textarea.value = text;
      textarea.setAttribute('readonly', '');
      textarea.style.position = 'fixed';
      textarea.style.left = '-9999px';
      document.body.appendChild(textarea);
      textarea.select();
      document.execCommand('copy');
      textarea.remove();
    }

    function exportFilename() {
      const ico = input.value.trim() || 'export';
      return `ico-scraper-${ico}.json`;
    }

    function downloadJson() {
      const blob = new Blob([latestRawJson], { type: 'application/json;charset=utf-8' });
      const url = URL.createObjectURL(blob);
      const link = document.createElement('a');
      link.href = url;
      link.download = exportFilename();
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    }

    async function downloadXlsx() {
      if (!latestData) return;

      const response = await fetch('/export/xlsx', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(latestData)
      });

      if (!response.ok) {
        const error = await response.json().catch(() => ({}));
        throw new Error(error.detail || 'Export do XLSX zlyhal.');
      }

      const blob = await response.blob();
      const url = URL.createObjectURL(blob);
      const link = document.createElement('a');
      link.href = url;
      link.download = exportFilename().replace(/\\.json$/, '.xlsx');
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    }

    function bindExportButtons() {
      const downloadButton = document.querySelector('#download-export');
      const downloadXlsxButton = document.querySelector('#download-xlsx');
      const copyButton = document.querySelector('#copy-export');

      if (downloadButton) {
        downloadButton.addEventListener('click', downloadJson);
      }

      if (downloadXlsxButton) {
        downloadXlsxButton.addEventListener('click', async () => {
          try {
            await downloadXlsx();
          } catch (error) {
            statusEl.className = 'status error';
            statusEl.textContent = error.message;
          }
        });
      }

      if (copyButton) {
        copyButton.addEventListener('click', async () => {
          try {
            await copyText(latestRawJson);
            const originalHtml = copyButton.innerHTML;
            copyButton.textContent = 'Skopírované';
            setTimeout(() => { copyButton.innerHTML = originalHtml; }, 1600);
          } catch (error) {
            statusEl.className = 'status error';
            statusEl.textContent = 'Kopírovanie zlyhalo.';
          }
        });
      }
    }




    function renderCrz(value) {
      if (!value || value.message) {
        return renderValue(value);
      }

      const contracts = Array.isArray(value.zmluvy) ? value.zmluvy : [];
      if (contracts.length === 0) {
        return '<div class="notice">V CRZ sa nepodarilo načítať zmluvy.</div>';
      }

      return `
        <div class="data-table-wrap">
          <table class="data-table">
            <thead>
              <tr>
                <th>Dátum</th>
                <th>Zmluva</th>
                <th>Cena</th>
                <th>Dodávateľ</th>
                <th>Odberateľ</th>
              </tr>
            </thead>
            <tbody>
              ${contracts.map(contract => `
                <tr>
                  <td>${renderValue(contract.datum || '-')}</td>
                  <td>
                    ${contract.link ? `<a class="document-link" href="${escapeAttribute(contract.link)}" target="_blank" rel="noopener noreferrer">${renderValue(contract.nazov_zmluvy || '-')}</a>` : renderValue(contract.nazov_zmluvy || '-')}
                    ${contract.cislo_zmluvy ? `<br><span>${renderValue(contract.cislo_zmluvy)}</span>` : ''}
                  </td>
                  <td>${renderValue(contract.cena || '-')}</td>
                  <td>${renderValue(contract.dodavatel || '-')}</td>
                  <td>${renderValue(contract.odberatel || '-')}</td>
                </tr>
              `).join('')}
            </tbody>
          </table>
        </div>
      `;
    }

    function renderRuzRecordsTable(records, tableId) {
      if (!Array.isArray(records) || records.length === 0) {
        return '<div class="notice">Záznamy sa nepodarilo načítať.</div>';
      }

      const typeColumn = records.some(record => Object.prototype.hasOwnProperty.call(record, 'Typ výročnej správy'))
        ? 'Typ výročnej správy'
        : 'Typ závierky';
      const columns = ['Obdobie', typeColumn, 'Predložená dňa', 'Zostavená dňa', 'Schválená dňa', 'SA uložená dňa'];
      return `
        <div class="data-table-wrap">
          <table class="data-table">
            <thead>
              <tr>${columns.map(column => `<th>${renderValue(column)}</th>`).join('')}</tr>
            </thead>
            <tbody>
              ${records.map((record, index) => `
                <tr>
                  ${columns.map((column, columnIndex) => `
                    <td>
                      ${columnIndex === 0
                        ? `<button class="period-button" type="button" data-documents-toggle="${tableId}-${index}" aria-expanded="false">${renderValue(record[column] || '-')}</button>`
                        : renderValue(record[column] || '-')}
                    </td>
                  `).join('')}
                </tr>
                <tr class="documents-row" data-documents-row="${tableId}-${index}">
                  <td class="documents-cell" colspan="${columns.length}">
                    ${renderDocumentsTable(record.dokumenty || [])}
                  </td>
                </tr>
              `).join('')}
            </tbody>
          </table>
        </div>
      `;
    }

    function renderDocumentsTable(documents) {
      if (!Array.isArray(documents) || documents.length === 0) {
        return '<div class="notice">Pre tento záznam nie sú dostupné dokumenty.</div>';
      }

      return `
        <p class="documents-title">Dokumenty</p>
        <div class="data-table-wrap">
          <table class="data-table">
            <thead>
              <tr>
                <th>Názov</th>
                <th>Typ</th>
                <th>Zdroj</th>
              </tr>
            </thead>
            <tbody>
              ${documents.map(documentItem => `
                <tr>
                  <td>${documentItem.url ? `<a class="document-link" href="${escapeAttribute(documentItem.url)}" target="_blank" rel="noopener noreferrer">${renderValue(documentItem.nazov || '-')}</a>` : renderValue(documentItem.nazov || '-')}</td>
                  <td>${renderValue(documentItem.typ || '-')}</td>
                  <td>${renderValue(documentItem.zdroj || '-')}</td>
                </tr>
              `).join('')}
            </tbody>
          </table>
        </div>
      `;
    }

    function bindRuzTables() {
      document.querySelectorAll('[data-documents-toggle]').forEach(toggle => {
        toggle.addEventListener('click', () => {
          const row = document.querySelector(`[data-documents-row="${toggle.dataset.documentsToggle}"]`);
          if (!row) return;
          const isVisible = row.classList.toggle('visible');
          toggle.setAttribute('aria-expanded', isVisible ? 'true' : 'false');
        });
      });
    }

    function renderRuz(value) {
      if (!value || value.message) {
        return renderValue(value);
      }

      const statements = Array.isArray(value.uctovne_zavierky)
        ? value.uctovne_zavierky
        : Array.isArray(value.zaznamy) ? value.zaznamy : [];
      const annualReports = Array.isArray(value.vyrocne_spravy) ? value.vyrocne_spravy : [];
      const rest = Object.fromEntries(
        Object.entries(value).filter(([key]) => !['zaznamy', 'uctovne_zavierky', 'vyrocne_spravy'].includes(key))
      );

      return `
        ${renderValue(rest)}
        <h3 class="section-subtitle">Účtovné závierky</h3>
        ${renderRuzRecordsTable(statements, 'statements')}
        <h3 class="section-subtitle">Výročné správy</h3>
        ${renderRuzRecordsTable(annualReports, 'annual-reports')}
      `;
    }

    function renderPanel(key, value, active) {
      if (key === 'raw') {
        return `
          <section class="panel ${active ? 'active' : ''}" id="panel-${key}" role="tabpanel">
            <h2 class="panel-title">Export dát</h2>
            <div class="export-panel">
              <p>Stiahnite si kompletný výsledok vo formáte JSON.</p>
              <div class="export-actions">
                <button class="export-button" id="download-export" type="button">${downloadIcon} Stiahnuť .json</button>
                <button class="export-button" id="download-xlsx" type="button">${downloadIcon} Stiahnuť .xlsx</button>
                <button class="export-button secondary" id="copy-export" type="button">${copyIcon} Kopírovať JSON</button>
              </div>
            </div>
          </section>
        `;
      }

      if (key === 'ico') {
        return `
          <section class="panel ${active ? 'active' : ''}" id="panel-${key}" role="tabpanel">
            <h2 class="panel-title">Všeobecné Informácie</h2>
            ${renderSummary(value)}
          </section>
        `;
      }

      const displayValue = key === 'finstat' && value && typeof value === 'object'
        ? Object.fromEntries(Object.entries(value).filter(([childKey]) => childKey !== 'grafy'))
        : value;
      const body = key === 'ruz'
        ? renderRuz(displayValue)
        : key === 'crz' ? renderCrz(displayValue) : renderValue(displayValue);

      return `
        <section class="panel ${active ? 'active' : ''}" id="panel-${key}" role="tabpanel">
          <h2 class="panel-title">${formatKey(key)}</h2>
          ${body}
        </section>
      `;
    }

    function activateTab(key) {
      document.querySelectorAll('.tab-button').forEach(tab => {
        tab.classList.toggle('active', tab.dataset.tab === key);
        tab.setAttribute('aria-selected', tab.dataset.tab === key ? 'true' : 'false');
      });
      document.querySelectorAll('.panel').forEach(panel => {
        panel.classList.toggle('active', panel.id === `panel-${key}`);
      });
    }

    function renderResults(data) {
      latestData = data;
      latestRawJson = JSON.stringify(data, null, 2);
      const includeGeneralInfo = sourceInputs.every(input => input.checked);
      const preferredOrder = [
        ...(includeGeneralInfo ? ['ico'] : []),
        'orsr',
        'rpvs',
        'finstat',
        'ruz',
        'crz'
      ];
      const keys = preferredOrder.filter(key => key in data);
      Object.keys(data).forEach(key => {
        if (key === 'ico' && !includeGeneralInfo) return;
        if (!keys.includes(key)) keys.push(key);
      });
      keys.push('raw');

      tabs.className = 'tabs visible';
      tabs.innerHTML = keys.map((key, index) => `
        <button class="tab-button ${index === 0 ? 'active' : ''}" type="button" data-tab="${key}" aria-controls="panel-${key}" aria-selected="${index === 0 ? 'true' : 'false'}">
          ${formatKey(key)}
        </button>
      `).join('');

      results.innerHTML = keys.map((key, index) => {
        const value = key === 'raw' || key === 'ico' ? data : data[key];
        return renderPanel(key, value, index === 0);
      }).join('');

      tabs.querySelectorAll('.tab-button').forEach(tab => {
        tab.addEventListener('click', () => activateTab(tab.dataset.tab));
      });
      bindExportButtons();
      bindRuzTables();
    }

    form.addEventListener('submit', async (event) => {
      event.preventDefault();
      const ico = input.value.trim();
      const sources = selectedSources();
      if (sources.length === 0) {
        statusEl.className = 'status error';
        statusEl.textContent = 'Vyberte aspoň jeden zdroj informácií.';
        clearResults();
        return;
      }
      statusEl.className = 'status';
      statusEl.textContent = 'Spracúvam požiadavku...';
      button.disabled = true;
      clearResults('Čakám na výsledok zo zdrojov...');

      try {
        const params = new URLSearchParams({ ico, sources: sources.join(',') });
        const response = await fetch(`/scrape?${params.toString()}`);
        const data = await response.json();
        if (!response.ok) {
          throw new Error(data.detail || 'Požiadavka zlyhala.');
        }
        renderResults(data);
        statusEl.textContent = 'Hotovo.';
      } catch (error) {
        statusEl.className = 'status error';
        statusEl.textContent = error.message;
        clearResults('Výsledok sa nepodarilo načítať.');
      } finally {
        button.disabled = false;
      }
    });
  </script>
</body>
</html>
"""


def is_base64_image(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("data:image/")


def flatten_value(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []

    if is_base64_image(value):
        return [(prefix or "image", "[image omitted]")]

    if isinstance(value, dict):
        if not value:
            return [(prefix, "")]
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(flatten_value(child, child_prefix))
        return rows

    if isinstance(value, list):
        if not value:
            return [(prefix, "")]
        for index, item in enumerate(value, start=1):
            child_prefix = f"{prefix}[{index}]" if prefix else f"[{index}]"
            rows.extend(flatten_value(item, child_prefix))
        return rows

    return [(prefix, "" if value is None else str(value))]


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)[:31].strip()
    return cleaned or "Sheet"


def build_xlsx(data: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="E4E7EB")
    headers = ("Pole", "Hodnota")

    for section, value in data.items():
        worksheet = workbook.create_sheet(safe_sheet_name(str(section)))
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for field, field_value in flatten_value(value):
            worksheet.append((field, field_value))

        worksheet.column_dimensions["A"].width = 42
        worksheet.column_dimensions["B"].width = 90

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def normalize_ico(ico: str) -> str:
    normalized = re.sub(r"\s+", "", ico)
    if not re.fullmatch(r"\d{8}", normalized):
        raise HTTPException(status_code=400, detail="IČO musí obsahovať presne 8 číslic.")
    return normalized


def normalize_sources(sources: str | None) -> set[str]:
    if not sources:
        return set(AVAILABLE_SOURCES)

    selected = {source.strip().lower() for source in sources.split(",") if source.strip()}
    invalid = selected - AVAILABLE_SOURCES
    if invalid:
        raise HTTPException(status_code=400, detail=f"Neplatné zdroje: {', '.join(sorted(invalid))}.")
    if not selected:
        raise HTTPException(status_code=400, detail="Vyberte aspoň jeden zdroj informácií.")
    return selected


@app.get("/", response_class=HTMLResponse)
async def index() -> str:
    return INDEX_HTML


@app.get("/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/export/xlsx")
async def export_xlsx(request: Request) -> StreamingResponse:
    try:
        data = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Neplatné JSON dáta pre export.") from exc

    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="Export očakáva JSON objekt.")

    output = await run_in_threadpool(build_xlsx, data)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ico-scraper-export.xlsx"'},
    )


@app.get("/scrape")
async def scrape(
    ico: str = Query(..., description="Slovak company IČO"),
    sources: str | None = Query(None, description="Comma-separated source keys"),
) -> dict:
    normalized_ico = normalize_ico(ico)
    selected_sources = normalize_sources(sources)

    async with _scrape_limit:
        try:
            return await run_in_threadpool(scrape_subject, normalized_ico, selected_sources)
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Scraping zlyhal: {type(exc).__name__}: {exc}") from exc
